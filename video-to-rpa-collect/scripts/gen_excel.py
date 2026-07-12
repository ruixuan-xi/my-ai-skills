# -*- coding: utf-8 -*-
"""
Generate 需求收集表 Excel from a JSON config.

Config schema (JSON):
{
    "template_path": ".../需求收集模板.xlsx",
    "output_path": ".../需求收集表_YYYYMMDD.xlsx",
    "generated_at": "auto",
    "items": [
        {
            "video_file": "运营_延慧美_设置各类促销.mp4",
            "department": "运营",              // or null if 待补充
            "submitter": "延慧美",              // or null
            "workflow_name": "设置各类促销",     // or null
            "steps": "1. ...\\n2. ...",          // string, newline-separated number list
            "pain_points": "...",               // or null
            "systems": "京麦",                  // or null
            "enterprise": "北京纳米科技",        // or null
            "missing_fields": ["部门", "需求提交人", ...]
        }
    ]
}

Missing fields are filled with "⚠待补充" and the cell is highlighted yellow.
Each item starts a new row; the 示例 row in the template (row 2) is replaced
so that the first data row is row 2.
"""
import json
import os
import sys
import shutil
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


MISSING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")   # light blue
THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

HEADERS = ["部门", "需求提交人", "工作流程名称", "具体工作步骤说明", "业务痛点", "涉及到的软件系统和网页", "企业名称"]
FIELD_KEYS = ["department", "submitter", "workflow_name", "steps", "pain_points", "systems", "enterprise"]


def parse_filename(filename: str):
    """Parse filename like '部门_提交人_流程名.mp4' into dict of possibly-null fields."""
    base = os.path.splitext(os.path.basename(filename))[0]
    parts = base.split("_")
    parts = [p.strip() for p in parts if p.strip()]

    result = {
        "department": None,
        "submitter": None,
        "workflow_name": None,
        "missing_fields": [],
    }

    # Heuristic:
    # - If 3+ segments: treat first as dept, second as submitter, rest joined as workflow_name
    # - If 2 segments: likely dept + workflow_name (missing submitter)
    # - If 1 segment: workflow_name only (dept and submitter missing)
    if len(parts) >= 3:
        result["department"] = parts[0]
        result["submitter"] = parts[1]
        result["workflow_name"] = "_".join(parts[2:])
    elif len(parts) == 2:
        result["department"] = parts[0]
        result["workflow_name"] = parts[1]
        result["missing_fields"].append("需求提交人")
    else:
        result["workflow_name"] = parts[0] if parts else None
        result["missing_fields"].append("部门")
        result["missing_fields"].append("需求提交人")
        if not result["workflow_name"]:
            result["missing_fields"].append("工作流程名称")

    if not result["department"]:
        if "部门" not in result["missing_fields"]:
            result["missing_fields"].append("部门")
    if not result["submitter"]:
        if "需求提交人" not in result["missing_fields"]:
            result["missing_fields"].append("需求提交人")
    if not result["workflow_name"]:
        if "工作流程名称" not in result["missing_fields"]:
            result["missing_fields"].append("工作流程名称")

    return result


def scan_videos(folder):
    """Scan folder (recursively) for video files, parse filenames into skeleton items."""
    VIDEO_EXTS = {".mp4", ".mov", ".avi", ".mkv", ".wmv", ".flv", ".webm", ".m4v"}
    items = []
    for root, _dirs, files in os.walk(folder):
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in VIDEO_EXTS:
                full = os.path.join(root, f).replace("\\", "/")
                parsed = parse_filename(f)
                items.append({
                    "video_file": full,
                    "video_filename": f,
                    **parsed,
                    "steps": None,
                    "pain_points": None,
                    "systems": None,
                    "enterprise": None,
                })
    # add missing_fields for fields that AI will fill in later (steps/systems/... are always待补充 initially)
    for it in items:
        for label, key in [("具体工作步骤说明", "steps"), ("业务痛点", "pain_points"),
                           ("涉及到的软件系统和网页", "systems"), ("企业名称", "enterprise")]:
            if not it.get(key):
                if label not in it["missing_fields"]:
                    it["missing_fields"].append(label)
    return items


def _cell_value(item, key):
    """Return display value; None becomes '⚠待补充'."""
    v = item.get(key)
    if v is None or (isinstance(v, str) and not v.strip()):
        return "⚠待补充"
    return v


def generate(config_path):
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    template_path = config["template_path"]
    output_path = config["output_path"]
    items = config.get("items", [])

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active  # 需求收集表 sheet

    # Delete all data rows except the header (row 1)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Re-assert header style (in case template formatting was lost)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c, value=HEADERS[c - 1])
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.font = Font(bold=True, name="微软雅黑", size=10)
        cell.border = THIN_BORDER

    # Auto-fit row heights for wrapped text
    for idx, item in enumerate(items, start=2):
        # compute missing fields set for quick lookup
        missing_set = set(item.get("missing_fields", []))

        for col_idx, (header_label, key) in enumerate(zip(HEADERS, FIELD_KEYS), start=1):
            value = _cell_value(item, key)
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGN
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = THIN_BORDER

            # yellow highlight for missing fields
            if header_label in missing_set or (key in item and (item.get(key) is None or (isinstance(item.get(key), str) and not item.get(key).strip()))):
                cell.fill = MISSING_FILL

        # Estimate row height based on longest multi-line cell (particularly steps col D)
        steps_val = item.get("steps") or ""
        line_count = steps_val.count("\n") + 1 if steps_val else 1
        # Roughly 18px per line; minimum 30px
        ws.row_dimensions[idx].height = max(30, min(line_count * 22, 400))

    # Ensure column widths are set from template or defaults
    default_widths = {"A": 10, "B": 14, "C": 22, "D": 60, "E": 22, "F": 28, "G": 18}
    from openpyxl.utils import get_column_letter
    for c, w in default_widths.items():
        if ws.column_dimensions[c].width is None or ws.column_dimensions[c].width < 10:
            ws.column_dimensions[c].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)

    n_items = len(items)
    n_missing_dept = sum(1 for it in items if "部门" in it.get("missing_fields", []))
    n_missing_sub = sum(1 for it in items if "需求提交人" in it.get("missing_fields", []))
    print(f"OK: 需求收集表已生成 -> {output_path}")
    print(f"共 {n_items} 条需求")
    print(f"  - 缺失「部门」: {n_missing_dept} 条")
    print(f"  - 缺失「需求提交人」: {n_missing_sub} 条")
    print(f"  - 待补「步骤/痛点/系统/企业」字段会标黄色⚠待补充，请让用户或AI补充完整")
    return output_path


def cmd_scan(args):
    """Subcommand: scan a folder and emit a JSON skeleton to stdout/file."""
    folder = args.folder
    if not os.path.isdir(folder):
        print(f"ERROR: not a directory: {folder}", file=sys.stderr)
        sys.exit(1)
    items = scan_videos(folder)
    # add placeholder for non-parsed fields
    skeleton = {
        "template_path": args.template or os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "references", "需求收集模板.xlsx"),
        "output_path": args.output or os.path.join(folder, f"需求收集表_{datetime.now().strftime('%Y%m%d')}.xlsx"),
        "items": items,
    }
    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(skeleton, f, ensure_ascii=False, indent=2)
        print(f"JSON skeleton written to: {args.json_out}")
        print(f"Found {len(items)} video(s).")
    else:
        print(json.dumps(skeleton, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser(description="Generate 需求收集表 Excel from config JSON, or scan folder for skeleton")
    parser.add_argument("--template", help="(scan mode) path to template xlsx")
    parser.add_argument("--output", help="output xlsx path")
    parser.add_argument("config_or_cmd", nargs="?", help="Path to config JSON, or subcommand 'scan'")
    parser.add_argument("folder", nargs="?", help="(scan mode) Folder containing videos")
    parser.add_argument("--json-out", dest="json_out", help="(scan mode) Write JSON skeleton to this path")

    args = parser.parse_args()

    if args.config_or_cmd == "scan":
        # re-construct scan args namespace
        class ScanArgs:
            folder = args.folder
            template = args.template
            output = args.output
            json_out = args.json_out
        if not ScanArgs.folder:
            parser.error("scan mode requires a folder argument")
        cmd_scan(ScanArgs())
        return

    if not args.config_or_cmd:
        parser.print_help()
        sys.exit(1)

    generate(args.config_or_cmd)


if __name__ == "__main__":
    main()
